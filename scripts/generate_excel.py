#!/usr/bin/env python3
"""
Génère un export Excel formaté avec plusieurs feuilles
Feuilles: Overview, Detailed, By Manufacturer, By SoC
"""

import json
import sys
from pathlib import Path
from collections import OrderedDict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd
except ImportError:
    print("❌ openpyxl/pandas not installed. Run: pip install openpyxl pandas")
    sys.exit(1)

def load_json(filepath):
    """Charge un fichier JSON"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)

def flatten_board(board):
    """Convertit un board JSON en dictionnaire plat"""
    flat = OrderedDict()
    flat['ID'] = board.get('id', '')
    flat['Manufacturer'] = board.get('manufacturer', '')
    flat['Reference'] = board.get('reference', '')
    flat['SoC'] = board.get('soc', {}).get('name', '')
    flat['Cores'] = board.get('soc', {}).get('cores', '')
    flat['Frequency'] = f"{board.get('soc', {}).get('frequency_mhz', '')} MHz"
    flat['Flash'] = f"{board.get('memory', {}).get('flash_mb', '')} MB"
    flat['PSRAM'] = f"{board.get('memory', {}).get('psram_mb', '')} MB"
    flat['Display'] = 'Yes' if board.get('display', {}).get('enabled') else 'No'
    flat['Display Size'] = board.get('display', {}).get('size_inches', '')
    flat['Wi-Fi 6'] = 'Yes' if board.get('connectivity', {}).get('wifi6') else 'No'
    flat['Zigbee'] = 'Yes' if board.get('connectivity', {}).get('zigbee') else 'No'
    flat['Matter'] = 'Yes' if board.get('connectivity', {}).get('matter') else 'No'
    flat['Price USD'] = board.get('price_usd', '')
    flat['Status'] = board.get('status', '')
    return flat

def style_header(ws):
    """Style l'en-tête"""
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def auto_adjust_columns(ws):
    """Ajuste automatiquement la largeur des colonnes"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    script_dir = Path(__file__).parent.parent
    boards_file = script_dir / "database" / "boards.json"
    excel_file = script_dir / "database" / "boards.xlsx"
    
    print("📊 Génération du fichier Excel...")
    print(f"📁 Source: {boards_file}")
    print(f"💾 Destination: {excel_file}\n")
    
    try:
        boards = load_json(boards_file)
    except Exception as e:
        print(f"❌ Erreur lors du chargement: {e}")
        return 1
    
    if not boards:
        print("⚠️  Aucune carte trouvée")
        return 1
    
    try:
        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut
        
        # === Feuille 1: Vue d'ensemble (Overview) ===
        flat_boards = [flatten_board(b) for b in boards]
        df_overview = pd.DataFrame(flat_boards)
        
        ws_overview = wb.create_sheet("Overview")
        for r_idx, row in enumerate(df_overview.values, 1):
            for c_idx, value in enumerate(row, 1):
                ws_overview.cell(row=r_idx, column=c_idx, value=value)
        
        # Ajouter l'en-tête
        for c_idx, col in enumerate(df_overview.columns, 1):
            ws_overview.cell(row=1, column=c_idx, value=col)
        
        style_header(ws_overview)
        ws_overview.auto_filter.ref = f"A1:{get_column_letter(len(df_overview.columns))}" + str(len(df_overview) + 1)
        auto_adjust_columns(ws_overview)
        
        # === Feuille 2: Groupé par Fabricant ===
        manufacturers = {}
        for board in boards:
            mfg = board.get('manufacturer', 'Unknown')
            if mfg not in manufacturers:
                manufacturers[mfg] = []
            manufacturers[mfg].append(board)
        
        ws_mfg = wb.create_sheet("By Manufacturer")
        row = 1
        
        for mfg in sorted(manufacturers.keys()):
            # Titre du fabricant
            ws_mfg.merge_cells(f'A{row}:E{row}')
            cell = ws_mfg[f'A{row}']
            cell.value = f"🏭 {mfg} ({len(manufacturers[mfg])} boards)"
            cell.font = Font(bold=True, size=12, color="1F4E78")
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1
            
            # Boards du fabricant
            for board in manufacturers[mfg]:
                flat = flatten_board(board)
                ws_mfg[f'A{row}'] = flat['Reference']
                ws_mfg[f'B{row}'] = flat['SoC']
                ws_mfg[f'C{row}'] = flat['Display']
                ws_mfg[f'D{row}'] = flat['Price USD']
                ws_mfg[f'E{row}'] = flat['Status']
                row += 1
            
            row += 1  # Espace
        
        ws_mfg.column_dimensions['A'].width = 30
        ws_mfg.column_dimensions['B'].width = 15
        ws_mfg.column_dimensions['C'].width = 10
        ws_mfg.column_dimensions['D'].width = 12
        ws_mfg.column_dimensions['E'].width = 12
        
        # === Feuille 3: Groupé par SoC ===
        socs = {}
        for board in boards:
            soc = board.get('soc', {}).get('name', 'Unknown')
            if soc not in socs:
                socs[soc] = []
            socs[soc].append(board)
        
        ws_soc = wb.create_sheet("By SoC")
        row = 1
        
        for soc in sorted(socs.keys()):
            ws_soc.merge_cells(f'A{row}:C{row}')
            cell = ws_soc[f'A{row}']
            cell.value = f"🔧 {soc} ({len(socs[soc])} boards)"
            cell.font = Font(bold=True, size=12, color="1F4E78")
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1
            
            for board in socs[soc]:
                flat = flatten_board(board)
                ws_soc[f'A{row}'] = flat['Manufacturer']
                ws_soc[f'B{row}'] = flat['Reference']
                ws_soc[f'C{row}'] = flat['Price USD']
                row += 1
            
            row += 1
        
        ws_soc.column_dimensions['A'].width = 20
        ws_soc.column_dimensions['B'].width = 35
        ws_soc.column_dimensions['C'].width = 12
        
        # Sauvegarder
        wb.save(excel_file)
        
        print(f"✅ Excel généré avec succès!")
        print(f"📍 Fichier: {excel_file.name}")
        print(f"📄 Feuilles: Overview, By Manufacturer, By SoC")
        print(f"📊 Total: {len(boards)} cartes")
        print("\n✅ Export réussi!")
        return 0
        
    except Exception as e:
        print(f"❌ Erreur lors de la génération: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
